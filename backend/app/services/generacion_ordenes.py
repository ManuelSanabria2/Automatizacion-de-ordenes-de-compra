"""Generación de la Orden de Compra oficial (INSTRUCCIONS.MD §6 y §7 paso 6-7).

Reglas invariables que implementa este módulo:
- Consecutivo SOLO vía la RPC `siguiente_numero_orden` (nunca calculado aquí).
- Orden de cálculo (§6.5): descuento por ítem → subtotal → descuento general
  → base IVA → IVA → total. Los valores se escriben como números calculados
  (nunca fórmulas ni digitación manual): documento, base de datos y respuesta
  de la API comparten exactamente las mismas cifras.
- `descripcion_final` se CONGELA con el nombre oficial del catálogo al momento
  de generar; cambios posteriores del catálogo no alteran órdenes emitidas.
- Las secciones fijas de la plantilla (§6.1) nunca se modifican: el documento
  se construye sobre una copia de la plantilla oficial y solo se escriben las
  celdas variables. Nota openpyxl 3.1.5: insert_rows/delete_rows desplaza
  valores, fórmulas y estilos pero NO celdas combinadas ni altos de fila —
  `_ajustar_filas` los desplaza manualmente.
"""

import re
from copy import copy
from datetime import date
from io import BytesIO

import openpyxl
from num2words import num2words
from pydantic import BaseModel, Field, field_validator
from storage3.exceptions import StorageApiError

from app.core import config
from app.core.supabase import obtener_cliente

# --- Layout de la plantilla oficial -------------------------------------------

FILA_PRIMER_ITEM = 15
FILAS_ITEMS_PLANTILLA = 5  # la plantilla trae 5 filas de ítems (15-19)
COLUMNAS_TABLA = "BCDEFG"
FORMATO_PORCENTAJE = "0%"

MESES = [
    "enero", "febrero", "marzo", "abril", "mayo", "junio",
    "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre",
]

BUCKET_ORDENES = "ordenes"
BUCKET_COTIZACIONES = "cotizaciones"
MIME_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# --- Errores de dominio --------------------------------------------------------


class ErrorProductoInexistente(ValueError):
    """Un ítem referencia un producto que no existe en el catálogo (400)."""


class ErrorOrdenNoEncontrada(ValueError):
    """No existe una orden con ese id (404)."""


# --- Modelos de entrada (espejo del orden_borrador del frontend) ---------------


class ProveedorBorrador(BaseModel):
    nombre: str
    nit: str
    direccion: str = ""
    ciudad: str = ""

    @field_validator("nombre", "nit")
    @classmethod
    def _no_vacio(cls, v: str) -> str:
        if not v.strip():
            raise ValueError("no puede estar vacío")
        return v.strip()


class ItemBorrador(BaseModel):
    item_num: int = Field(ge=1)
    producto_empresa_id: str
    descripcion_proveedor: str = ""
    unidad: str = ""
    cantidad: float = Field(gt=0)
    valor_unitario: float = Field(ge=0)
    descuento_porcentaje: float = Field(default=0, ge=0, le=100)


class OrdenBorrador(BaseModel):
    proveedor: ProveedorBorrador
    proveedor_existente: bool = False
    numero_cotizacion: str = ""
    proyecto: str
    plazo_entrega: str
    forma_pago: str
    sitio_entrega: str
    tasa_iva: float = Field(default=19, ge=0, le=100)
    descuento_general_porcentaje: float = Field(default=0, ge=0, le=100)
    items: list[ItemBorrador] = Field(min_length=1)

    @field_validator("proyecto", "plazo_entrega", "forma_pago", "sitio_entrega")
    @classmethod
    def _no_vacio(cls, v: str) -> str:
        if not v.strip():
            raise ValueError("es un campo obligatorio")
        return v.strip()


# --- Modelos de salida ----------------------------------------------------------


class TotalesOrden(BaseModel):
    subtotal: float
    valor_descuento: float
    base_iva: float
    iva: float
    total: float


class OrdenGenerada(BaseModel):
    id: str
    numero_orden: str
    fecha: str  # ISO
    totales: TotalesOrden
    documento_path: str
    pdf_cotizacion_path: str | None


class OrdenResumen(BaseModel):
    id: str
    numero_orden: str
    fecha: str
    proveedor_nit: str | None
    proveedor_nombre: str | None
    total: float | None


class UrlDocumento(BaseModel):
    url: str


# --- Cálculo de totales (§6.5 — el orden importa) -------------------------------


class _ItemCalculado(BaseModel):
    item: ItemBorrador
    descripcion_final: str
    unidad: str
    valor_total: float


def calcular_totales(
    items: list[_ItemCalculado], descuento_general: float, tasa_iva: float
) -> TotalesOrden:
    subtotal = round(sum(it.valor_total for it in items), 2)
    valor_descuento = round(subtotal * descuento_general / 100, 2)
    base_iva = round(subtotal - valor_descuento, 2)
    iva = round(base_iva * tasa_iva / 100, 2)
    total = round(base_iva + iva, 2)
    return TotalesOrden(
        subtotal=subtotal,
        valor_descuento=valor_descuento,
        base_iva=base_iva,
        iva=iva,
        total=total,
    )


def _valor_total_item(item: ItemBorrador) -> float:
    return round(
        item.cantidad * item.valor_unitario * (1 - item.descuento_porcentaje / 100), 2
    )


# --- Utilidades de texto ---------------------------------------------------------


def _formato_pesos(valor: float) -> str:
    entero = int(valor)
    centavos = round((valor - entero) * 100)
    if centavos == 100:
        entero, centavos = entero + 1, 0
    miles = f"{entero:,}".replace(",", ".")
    return f"${miles},{centavos:02d}" if centavos else f"${miles}"


def _valor_en_letras(valor: float) -> str:
    entero = int(valor)
    centavos = round((valor - entero) * 100)
    if centavos == 100:
        entero, centavos = entero + 1, 0
    letras = num2words(entero, lang="es")
    # Apócope que num2words omite: "ochenta y uno mil" → "ochenta y un mil"
    letras = re.sub(r"\buno mil\b", "un mil", letras)
    letras = re.sub(r"\bveintiuno mil\b", "veintiún mil", letras)
    if centavos:
        return f"{letras} pesos con {centavos:02d}/100 m/cte"
    return f"{letras} pesos m/cte"


def _fecha_larga(fecha: date) -> str:
    return f"Garagoa, {fecha.day} de {MESES[fecha.month - 1]} del {fecha.year}"


def _titulo_hoja(nombre_proveedor: str, numero_orden: str) -> str:
    # Límite de Excel: 31 caracteres. Se recorta el nombre del proveedor,
    # nunca el consecutivo.
    consecutivo = numero_orden.rsplit("-", 1)[-1]
    nombre = re.sub(r"[\\/*?:\[\]']", " ", nombre_proveedor)
    nombre = re.sub(r"\s+", " ", nombre).strip()
    nombre = nombre[: 31 - len(consecutivo) - 1].strip()
    return f"{nombre} {consecutivo}" if nombre else consecutivo


# --- Manipulación de la plantilla ------------------------------------------------


def _ajustar_filas(ws, en_fila: int, delta: int) -> None:
    """Inserta (delta>0) o elimina (delta<0) filas desplazando también las
    celdas combinadas y los altos de fila, que openpyxl no mueve solo."""
    if delta == 0:
        return
    afectados = [r.bounds for r in list(ws.merged_cells.ranges) if r.min_row >= en_fila]
    for c1, f1, c2, f2 in afectados:
        ws.unmerge_cells(start_row=f1, start_column=c1, end_row=f2, end_column=c2)
    if delta > 0:
        ws.insert_rows(en_fila, delta)
    else:
        ws.delete_rows(en_fila, -delta)
    for c1, f1, c2, f2 in afectados:
        ws.merge_cells(
            start_row=f1 + delta, start_column=c1, end_row=f2 + delta, end_column=c2
        )
    alturas = {
        f: dim.height
        for f, dim in ws.row_dimensions.items()
        if f >= en_fila and dim.height is not None
    }
    for f in alturas:
        del ws.row_dimensions[f]
    for f, alto in alturas.items():
        if delta < 0 and f < en_fila - delta:
            continue  # filas eliminadas: su alto desaparece
        ws.row_dimensions[f + delta].height = alto


def _copiar_estilo_fila(ws, fila_origen: int, fila_destino: int) -> None:
    for col in COLUMNAS_TABLA:
        ws[f"{col}{fila_destino}"]._style = copy(ws[f"{col}{fila_origen}"]._style)
    alto = ws.row_dimensions[fila_origen].height
    if alto is not None:
        ws.row_dimensions[fila_destino].height = alto


def generar_documento(
    borrador: OrdenBorrador,
    items: list[_ItemCalculado],
    totales: TotalesOrden,
    numero_orden: str,
    fecha: date,
) -> bytes:
    """Llena la plantilla oficial con los datos de la orden y devuelve el .xlsx."""
    wb = openpyxl.load_workbook(config.PLANTILLA_ORDEN)
    ws = wb.active
    ws.title = _titulo_hoja(borrador.proveedor.nombre, numero_orden)

    n = len(items)

    # Tabla de ítems: ajustar el número de filas preservando estilos.
    if n > FILAS_ITEMS_PLANTILLA:
        extra = n - FILAS_ITEMS_PLANTILLA
        ultima_plantilla = FILA_PRIMER_ITEM + FILAS_ITEMS_PLANTILLA - 1
        # Se inserta ANTES de la última fila de ítems (dentro del bloque, para
        # no atravesar merges) y se copia el estilo de una fila de ítem intacta.
        _ajustar_filas(ws, ultima_plantilla, extra)
        for f in range(ultima_plantilla, ultima_plantilla + extra):
            _copiar_estilo_fila(ws, FILA_PRIMER_ITEM + 1, f)
    elif n < FILAS_ITEMS_PLANTILLA:
        _ajustar_filas(ws, FILA_PRIMER_ITEM + n, n - FILAS_ITEMS_PLANTILLA)

    for i, it in enumerate(items):
        f = FILA_PRIMER_ITEM + i
        ws[f"B{f}"] = float(it.item.item_num)
        ws[f"C{f}"] = it.descripcion_final
        ws[f"D{f}"] = it.unidad
        ws[f"E{f}"] = it.item.cantidad
        ws[f"F{f}"] = it.item.valor_unitario
        ws[f"G{f}"] = it.valor_total

    fila_subtotal = FILA_PRIMER_ITEM + n
    fila_iva = fila_subtotal + 1

    # Descuento general: fila propia entre el subtotal y el IVA, solo si aplica.
    if borrador.descuento_general_porcentaje > 0:
        _ajustar_filas(ws, fila_iva, 1)
        _copiar_estilo_fila(ws, fila_iva + 1, fila_iva)  # estilo de la fila IVA
        ws.merge_cells(
            start_row=fila_iva, start_column=2, end_row=fila_iva, end_column=5
        )
        ws[f"B{fila_iva}"] = "DESCUENTO GENERAL"
        celda_pct = ws[f"F{fila_iva}"]
        celda_pct.value = borrador.descuento_general_porcentaje / 100
        celda_pct.number_format = FORMATO_PORCENTAJE
        ws[f"G{fila_iva}"] = -totales.valor_descuento
        fila_iva += 1

    fila_total = fila_iva + 1

    ws[f"G{fila_subtotal}"] = totales.subtotal
    ws[f"F{fila_iva}"] = borrador.tasa_iva / 100
    ws[f"G{fila_iva}"] = totales.iva
    ws[f"G{fila_total}"] = totales.total

    # Encabezado y campos variables (las secciones fijas no se tocan).
    ws["B4"] = f"ORDEN DE COMPRA Nº {numero_orden}"
    ws["B5"] = _fecha_larga(fecha)
    ws["C6"] = borrador.proveedor.nombre
    ws["C7"] = borrador.proveedor.nit
    ws["C8"] = borrador.proveedor.direccion
    ws["C9"] = borrador.proveedor.ciudad
    ws["C10"] = borrador.proyecto

    objeto = (
        "Suministrar a ENERCER S.A. E.S.P los elementos relacionados a "
        "continuación, teniendo en cuenta las condiciones y especificaciones "
        "técnicas estipuladas en la presente Orden de Compra"
    )
    if borrador.numero_cotizacion:
        objeto += f" y la cotización No. {borrador.numero_cotizacion}"
    # B13 está antes de la tabla: no se desplaza.
    ws["B13"] = objeto

    # Bloques bajo la tabla: sus filas se desplazaron con la tabla.
    desplazamiento = (fila_total) - 22  # la fila del total en la plantilla es 22
    ws[f"B{25 + desplazamiento}"] = borrador.plazo_entrega
    ws[f"B{28 + desplazamiento}"] = (
        f"La presente orden tiene un valor de {_valor_en_letras(totales.total)}"
        f"({_formato_pesos(totales.total)}) Que serán cancelados "
        f"{borrador.forma_pago}."
    )
    ws[f"B{34 + desplazamiento}"] = (
        "Los elementos contenidos en la presente Orden de Compra se entregarán "
        f"en los siguientes sitios: {borrador.sitio_entrega}"
    )

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# --- Acceso a datos ---------------------------------------------------------------


def _productos_por_id(cliente, ids: list[str]) -> dict[str, dict]:
    unicos = list(dict.fromkeys(ids))
    respuesta = (
        cliente.table("productos_empresa")
        .select("id, nombre_oficial, unidad_default")
        .in_("id", unicos)
        .execute()
    )
    return {fila["id"]: fila for fila in respuesta.data}


def _preparar_items(cliente, borrador: OrdenBorrador) -> list[_ItemCalculado]:
    productos = _productos_por_id(
        cliente, [it.producto_empresa_id for it in borrador.items]
    )
    faltantes = [
        it.producto_empresa_id
        for it in borrador.items
        if it.producto_empresa_id not in productos
    ]
    if faltantes:
        raise ErrorProductoInexistente(
            f"Producto(s) inexistente(s) en el catálogo: {', '.join(faltantes)}"
        )
    calculados = []
    for it in borrador.items:
        producto = productos[it.producto_empresa_id]
        calculados.append(
            _ItemCalculado(
                item=it,
                # Nombre oficial CONGELADO al momento de la generación (§ integridad)
                descripcion_final=producto["nombre_oficial"],
                unidad=it.unidad.strip() or (producto["unidad_default"] or ""),
                valor_total=_valor_total_item(it),
            )
        )
    return calculados


def generar_orden(borrador: OrdenBorrador, pdf_cotizacion: bytes | None = None) -> OrdenGenerada:
    cliente = obtener_cliente()

    items = _preparar_items(cliente, borrador)
    totales = calcular_totales(
        items, borrador.descuento_general_porcentaje, borrador.tasa_iva
    )

    # Proveedor: si el NIT existe se actualiza, si no se crea (§6.3).
    cliente.table("proveedores").upsert(
        {
            "nit": borrador.proveedor.nit,
            "nombre": borrador.proveedor.nombre,
            "direccion": borrador.proveedor.direccion or None,
            "ciudad": borrador.proveedor.ciudad or None,
        },
        on_conflict="nit",
    ).execute()

    # Consecutivo atómico — única fuente permitida (§6.2).
    numero_orden = cliente.rpc("siguiente_numero_orden", {}).execute().data
    fecha = date.today()

    documento = generar_documento(borrador, items, totales, numero_orden, fecha)
    ruta_documento = f"{numero_orden}.xlsx"
    cliente.storage.from_(BUCKET_ORDENES).upload(
        ruta_documento, documento, {"content-type": MIME_XLSX}
    )

    ruta_pdf = None
    if pdf_cotizacion:
        ruta_pdf = f"{numero_orden}.pdf"
        cliente.storage.from_(BUCKET_COTIZACIONES).upload(
            ruta_pdf, pdf_cotizacion, {"content-type": "application/pdf"}
        )

    respuesta_orden = (
        cliente.table("ordenes_compra")
        .insert(
            {
                "numero_orden": numero_orden,
                "fecha": fecha.isoformat(),
                "proveedor_nit": borrador.proveedor.nit,
                "numero_cotizacion": borrador.numero_cotizacion or None,
                "proyecto": borrador.proyecto,
                "plazo_entrega": borrador.plazo_entrega,
                "forma_pago": borrador.forma_pago,
                "sitio_entrega": borrador.sitio_entrega,
                "tasa_iva": borrador.tasa_iva,
                "descuento_general_porcentaje": borrador.descuento_general_porcentaje,
                "subtotal": totales.subtotal,
                "valor_descuento": totales.valor_descuento,
                "iva": totales.iva,
                "total": totales.total,
                "pdf_cotizacion_url": f"{BUCKET_COTIZACIONES}/{ruta_pdf}" if ruta_pdf else None,
                "documento_generado_url": f"{BUCKET_ORDENES}/{ruta_documento}",
            }
        )
        .execute()
    )
    orden_id = respuesta_orden.data[0]["id"]

    try:
        cliente.table("ordenes_compra_items").insert(
            [
                {
                    "orden_id": orden_id,
                    "item_num": it.item.item_num,
                    "producto_empresa_id": it.item.producto_empresa_id,
                    "descripcion_final": it.descripcion_final,
                    "unidad": it.unidad or None,
                    "cantidad": it.item.cantidad,
                    "valor_unitario": it.item.valor_unitario,
                    "descuento_porcentaje": it.item.descuento_porcentaje,
                    "valor_total": it.valor_total,
                }
                for it in items
            ]
        ).execute()
    except Exception:
        # No dejar una orden sin ítems: se revierte el encabezado (cascade).
        cliente.table("ordenes_compra").delete().eq("id", orden_id).execute()
        raise

    return OrdenGenerada(
        id=orden_id,
        numero_orden=numero_orden,
        fecha=fecha.isoformat(),
        totales=totales,
        documento_path=f"{BUCKET_ORDENES}/{ruta_documento}",
        pdf_cotizacion_path=f"{BUCKET_COTIZACIONES}/{ruta_pdf}" if ruta_pdf else None,
    )


def listar_ordenes() -> list[OrdenResumen]:
    respuesta = (
        obtener_cliente()
        .table("ordenes_compra")
        .select("id, numero_orden, fecha, total, proveedor_nit, proveedores(nombre)")
        .order("created_at", desc=True)
        .execute()
    )
    return [
        OrdenResumen(
            id=fila["id"],
            numero_orden=fila["numero_orden"],
            fecha=fila["fecha"],
            proveedor_nit=fila["proveedor_nit"],
            proveedor_nombre=(fila.get("proveedores") or {}).get("nombre"),
            total=fila["total"],
        )
        for fila in respuesta.data
    ]


def url_documento(orden_id: str) -> UrlDocumento:
    """URL firmada (1 hora) para descargar el documento generado de una orden."""
    respuesta = (
        obtener_cliente()
        .table("ordenes_compra")
        .select("documento_generado_url")
        .eq("id", orden_id)
        .limit(1)
        .execute()
    )
    if not respuesta.data or not respuesta.data[0]["documento_generado_url"]:
        raise ErrorOrdenNoEncontrada(f"No existe documento para la orden {orden_id}")
    bucket, _, ruta = respuesta.data[0]["documento_generado_url"].partition("/")
    firmada = obtener_cliente().storage.from_(bucket).create_signed_url(ruta, 3600)
    url = firmada.get("signedURL") or firmada.get("signedUrl") or ""
    if not url:
        raise StorageApiError("No se pudo firmar la URL del documento", "", 502)
    return UrlDocumento(url=url)
